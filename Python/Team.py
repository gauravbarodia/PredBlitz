import xlsxwriter
import openpyxl

path1 = "test4.xlsx"

workbook_res = openpyxl.load_workbook(path1.strip())

result = workbook_res.active

Score = xlsxwriter.Workbook("TeamScore2.xlsx")
Fill = Score.add_worksheet("1")
Fill.write(0, 1, "Team")
Fill.write(0, 2, "Points")
teams = ['Chennai Super Kings (CSK)', 'Delhi Capitals (DC)', 'Gujarat Titans (GT)', 'Kolkata Knight Riders (KKR)',
         'Lucknow Super Giants (LSG)', 'Mumbai Indians (MI)', 'Punjab Kings (PBKS)', 'Rajasthan Royals (RR)',
         'Royal Challengers Bangalore (RCB)', 'Sunrisers Hyderabad (SRH)']
names = {team: 0 for team in teams}

for row in range(2, result.max_row + 1):
    key = result.cell(row=row, column=4).value
    val = result.cell(row=row, column=5).value
    names[key] += val


for i in range(0, 10):
    Fill.write(i + 1, 1, teams[i])
    Fill.write(i + 1, 2, names[teams[i]])


workbook_res.close()
Score.close()

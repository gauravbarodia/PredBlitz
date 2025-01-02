import xlsxwriter
import openpyxl

path1 ="KKRvsSRH (FINAL) (Responses).xlsx"
path2 = "Final_Result.xlsx"

workbook_resp = openpyxl.load_workbook(path1.strip())
workbook_res = openpyxl.load_workbook(path2.strip())

response = workbook_resp.active
result = workbook_res.active

Score = xlsxwriter.Workbook("test4.xlsx")
Fill = Score.add_worksheet("1")
Fill.write(0, 1, "Name")
Fill.write(0, 2, "Gmail")
Fill.write(0, 3, "Favourite Team")
Fill.write(0, 4, "Points")

for row in range(2, response.max_row + 1):
    s = 0
    for col in range(2, response.max_column+1):

        x = response.cell(row=row, column=col + 3).value

        y = result.cell(row=2, column=col).value

        if x is None or y is None:
            continue

        if col == 2:
            if x == y:
                s += 10
            continue
        elif (col > 2 and col < 7) or (col > 10 and col < 15):
            x = int(abs(x - y))
            if x < 10:
                s += 10 - x

        else:
            x = int(abs(x - y))
            if x == 0:
                s += 10
            elif x == 1:
                s += 5
            elif x == 2:
                s += 1
        
    Fill.write(row - 1, 1, response.cell(row=row, column=2).value)
    Fill.write(row - 1, 2, response.cell(row=row, column=3).value)
    Fill.write(row - 1, 3, response.cell(row=row, column=4).value)
    Fill.write(row - 1, 4, s)

workbook_resp.close()
workbook_res.close()
Score.close()
